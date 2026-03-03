"""Multi-format data exporter for PUC Rate Case Tracker.

Exports rate case data in CSV, JSON, Excel (multi-sheet), and Markdown
summary formats. Enriches rate case exports with utility details.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from rich.console import Console

console = Console()

PROJECT_ROOT = Path(__file__).parent.parent.parent
DEFAULT_EXPORT_DIR = PROJECT_ROOT / "data" / "exports"


def export_data(
    fmt: str = "all",
    output_dir: str = "data/exports",
) -> list[Path]:
    """Export rate case data to the specified format(s).

    Args:
        fmt: Format to export — "csv", "json", "excel", "markdown", or "all".
        output_dir: Output directory path (relative to project root or absolute).

    Returns:
        List of paths to exported files.
    """
    from src.storage.database import get_all_rate_cases, get_all_utilities, get_connection, get_stats

    # Resolve output directory
    if Path(output_dir).is_absolute():
        out_dir = Path(output_dir)
    else:
        out_dir = PROJECT_ROOT / output_dir

    out_dir.mkdir(parents=True, exist_ok=True)

    conn = get_connection()
    cases = get_all_rate_cases(limit=10000, conn=conn)
    utilities = get_all_utilities(conn=conn)
    stats = get_stats(conn=conn, print_output=False)
    conn.close()

    if not cases:
        console.print("[yellow]No rate cases to export.[/yellow]")
        return []

    exported_files = []
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d")

    if fmt in ("csv", "all"):
        path = _export_csv(cases, out_dir, timestamp)
        exported_files.append(path)

    if fmt in ("json", "all"):
        path = _export_json(cases, utilities, stats, out_dir, timestamp)
        exported_files.append(path)

    if fmt in ("excel", "all"):
        path = _export_excel(cases, utilities, stats, out_dir, timestamp)
        if path:
            exported_files.append(path)

    if fmt in ("markdown", "all"):
        path = _export_markdown(cases, stats, out_dir, timestamp)
        exported_files.append(path)

    console.print(f"\n[bold green]Exported {len(exported_files)} files to {out_dir}[/bold green]")
    for f in exported_files:
        console.print(f"  [dim]{f.name}[/dim]")

    return exported_files


# --- CSV Export ---


def _export_csv(
    cases: list[dict], out_dir: Path, timestamp: str
) -> Path:
    """Export rate cases to CSV."""
    filename = f"puc_rate_cases_{timestamp}.csv"
    filepath = out_dir / filename

    # Define column order
    columns = [
        "docket_number", "utility_name", "canonical_utility_name",
        "state", "source", "case_type", "utility_type", "status",
        "filing_date", "decision_date", "effective_date",
        "requested_revenue_change", "approved_revenue_change",
        "rate_base", "return_on_equity",
        "requested_rate_change_pct", "approved_rate_change_pct",
        "quality_score", "source_url", "description",
    ]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for case in cases:
            writer.writerow(case)

    console.print(f"[green]CSV exported: {filename} ({len(cases)} records)[/green]")
    return filepath


# --- JSON Export ---


def _export_json(
    cases: list[dict],
    utilities: list[dict],
    stats: dict,
    out_dir: Path,
    timestamp: str,
) -> Path:
    """Export rate cases, utilities, and stats to JSON."""
    filename = f"puc_rate_cases_{timestamp}.json"
    filepath = out_dir / filename

    export_data = {
        "metadata": {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "total_records": len(cases),
            "total_utilities": len(utilities),
            "format_version": "1.0",
            "project": "PUC Rate Case Tracker",
            "author": "Nathan Goldberg",
            "contact": "nathanmauricegoldberg@gmail.com",
        },
        "statistics": stats,
        "rate_cases": cases,
        "utilities": utilities,
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(export_data, f, indent=2, default=str)

    console.print(f"[green]JSON exported: {filename} ({len(cases)} records)[/green]")
    return filepath


# --- Excel Export ---


def _export_excel(
    cases: list[dict],
    utilities: list[dict],
    stats: dict,
    out_dir: Path,
    timestamp: str,
) -> Optional[Path]:
    """Export to multi-sheet Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print("[yellow]openpyxl not installed; skipping Excel export[/yellow]")
        return None

    filename = f"puc_rate_cases_{timestamp}.xlsx"
    filepath = out_dir / filename

    wb = openpyxl.Workbook()

    # --- Sheet 1: Rate Cases ---
    ws = wb.active
    ws.title = "Rate Cases"

    columns = [
        ("Docket Number", "docket_number", 18),
        ("Utility Name", "canonical_utility_name", 35),
        ("State", "state", 8),
        ("Case Type", "case_type", 22),
        ("Utility Type", "utility_type", 15),
        ("Status", "status", 12),
        ("Filing Date", "filing_date", 14),
        ("Decision Date", "decision_date", 14),
        ("Requested Revenue ($M)", "requested_revenue_change", 22),
        ("Approved Revenue ($M)", "approved_revenue_change", 22),
        ("Rate Base ($M)", "rate_base", 16),
        ("ROE (%)", "return_on_equity", 10),
        ("Requested Change (%)", "requested_rate_change_pct", 20),
        ("Approved Change (%)", "approved_rate_change_pct", 20),
        ("Quality Score", "quality_score", 14),
        ("Source", "source", 20),
        ("Source URL", "source_url", 40),
    ]

    # Header styling
    header_fill = PatternFill(start_color="0984E3", end_color="0984E3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, (header, _key, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, case in enumerate(cases, 2):
        for col_idx, (_header, key, _width) in enumerate(columns, 1):
            value = case.get(key)
            # Use canonical name, fall back to raw name
            if key == "canonical_utility_name" and not value:
                value = case.get("utility_name")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if key in ("requested_revenue_change", "approved_revenue_change", "rate_base"):
                if value is not None:
                    cell.number_format = '#,##0.0'
            elif key in ("return_on_equity", "requested_rate_change_pct", "approved_rate_change_pct"):
                if value is not None:
                    cell.number_format = '0.00'
            elif key == "quality_score":
                if value is not None:
                    cell.number_format = '0.000'

    # Freeze top row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Sheet 2: Utilities ---
    ws2 = wb.create_sheet("Utilities")
    util_cols = [
        ("Name", "name", 35),
        ("Canonical Name", "canonical_name", 35),
        ("State", "state", 8),
        ("Utility Type", "utility_type", 15),
        ("Ownership Type", "ownership_type", 18),
        ("Parent Company", "parent_company", 30),
    ]

    for col_idx, (header, _key, width) in enumerate(util_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, util in enumerate(utilities, 2):
        for col_idx, (_header, key, _width) in enumerate(util_cols, 1):
            ws2.cell(row=row_idx, column=col_idx, value=util.get(key))

    ws2.freeze_panes = "A2"

    # --- Sheet 3: Summary Statistics ---
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    summary_font = Font(bold=True, size=12)

    row = 1
    ws3.cell(row=row, column=1, value="PUC Rate Case Tracker - Summary Statistics").font = Font(
        bold=True, size=14
    )
    row += 1
    ws3.cell(row=row, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    row += 1
    ws3.cell(row=row, column=1, value=f"Built by Nathan Goldberg | nathanmauricegoldberg@gmail.com")
    row += 2

    ws3.cell(row=row, column=1, value="Total Rate Cases").font = summary_font
    ws3.cell(row=row, column=2, value=stats.get("total_rate_cases", 0))
    row += 1

    ws3.cell(row=row, column=1, value="Unique Utilities").font = summary_font
    ws3.cell(row=row, column=2, value=stats.get("unique_utilities", 0))
    row += 2

    # By state
    ws3.cell(row=row, column=1, value="Cases by State").font = summary_font
    row += 1
    for state, count in stats.get("by_state", {}).items():
        ws3.cell(row=row, column=1, value=f"  {state}")
        ws3.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws3.cell(row=row, column=1, value="Financial Summary").font = summary_font
    row += 1
    fin = stats.get("financial", {})
    ws3.cell(row=row, column=1, value="  Total Requested ($M)")
    ws3.cell(row=row, column=2, value=fin.get("total_requested_M", 0))
    row += 1
    ws3.cell(row=row, column=1, value="  Total Approved ($M)")
    ws3.cell(row=row, column=2, value=fin.get("total_approved_M", 0))
    row += 1
    ws3.cell(row=row, column=1, value="  Average ROE (%)")
    ws3.cell(row=row, column=2, value=fin.get("avg_roe_pct", 0))

    wb.save(filepath)
    console.print(f"[green]Excel exported: {filename} ({len(cases)} records, 3 sheets)[/green]")
    return filepath


# --- Markdown Export ---


def _export_markdown(
    cases: list[dict], stats: dict, out_dir: Path, timestamp: str
) -> Path:
    """Export summary statistics in Markdown format."""
    filename = f"puc_rate_cases_summary_{timestamp}.md"
    filepath = out_dir / filename

    lines = []
    lines.append("# PUC Rate Case Tracker - Data Summary")
    lines.append("")
    lines.append(f"*Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}*")
    lines.append(f"*Built by Nathan Goldberg | nathanmauricegoldberg@gmail.com*")
    lines.append("")

    # Overview
    lines.append("## Overview")
    lines.append("")
    lines.append(f"- **Total Rate Cases:** {stats.get('total_rate_cases', 0)}")
    lines.append(f"- **Unique Utilities:** {stats.get('unique_utilities', 0)}")
    lines.append(f"- **States Covered:** {len(stats.get('by_state', {}))}")

    date_range = stats.get("date_range", {})
    if date_range.get("earliest_filing"):
        lines.append(
            f"- **Date Range:** {date_range['earliest_filing']} to {date_range['latest_filing']}"
        )
    lines.append("")

    # By State
    lines.append("## Cases by State")
    lines.append("")
    lines.append("| State | Count |")
    lines.append("|-------|-------|")
    for state, count in sorted(stats.get("by_state", {}).items()):
        lines.append(f"| {state} | {count} |")
    lines.append("")

    # By Case Type
    lines.append("## Cases by Type")
    lines.append("")
    lines.append("| Case Type | Count |")
    lines.append("|-----------|-------|")
    for ctype, count in stats.get("by_case_type", {}).items():
        lines.append(f"| {ctype} | {count} |")
    lines.append("")

    # Financial Summary
    lines.append("## Financial Summary")
    lines.append("")
    fin = stats.get("financial", {})
    lines.append(f"- **Cases with Revenue Data:** {fin.get('cases_with_revenue_data', 0)}")
    lines.append(f"- **Total Revenue Requested:** ${fin.get('total_requested_M', 0):,.1f}M")
    lines.append(f"- **Total Revenue Approved:** ${fin.get('total_approved_M', 0):,.1f}M")
    lines.append(f"- **Average Requested:** ${fin.get('avg_requested_M', 0):,.1f}M")
    lines.append(f"- **Average Approved:** ${fin.get('avg_approved_M', 0):,.1f}M")
    lines.append(f"- **Average ROE:** {fin.get('avg_roe_pct', 0):.2f}%")
    lines.append("")

    # Quality
    lines.append("## Data Quality")
    lines.append("")
    qual = stats.get("quality", {})
    lines.append(f"- **Average Quality Score:** {qual.get('avg_score', 0):.3f}")
    lines.append(f"- **Min Score:** {qual.get('min_score', 0):.3f}")
    lines.append(f"- **Max Score:** {qual.get('max_score', 0):.3f}")
    lines.append(f"- **Above Threshold (>=0.6):** {qual.get('above_threshold', 0)}")
    lines.append("")

    # Top cases by revenue request
    lines.append("## Largest Rate Cases by Revenue Request")
    lines.append("")
    lines.append("| Docket | Utility | State | Requested ($M) | Approved ($M) | Status |")
    lines.append("|--------|---------|-------|-----------------|---------------|--------|")

    sorted_cases = sorted(
        [c for c in cases if c.get("requested_revenue_change")],
        key=lambda x: abs(x.get("requested_revenue_change", 0)),
        reverse=True,
    )

    for case in sorted_cases[:15]:
        docket = case.get("docket_number", "")
        utility = case.get("canonical_utility_name") or case.get("utility_name", "")
        state = case.get("state", "")
        requested = case.get("requested_revenue_change")
        approved = case.get("approved_revenue_change")
        status = case.get("status", "")

        req_str = f"${requested:,.1f}" if requested is not None else "N/A"
        app_str = f"${approved:,.1f}" if approved is not None else "Pending"

        lines.append(f"| {docket} | {utility[:30]} | {state} | {req_str} | {app_str} | {status} |")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("*Data sourced from state PUC docket systems (PA, CA, OR, IN, WA).*")
    lines.append("*Contact: nathanmauricegoldberg@gmail.com*")

    filepath.write_text("\n".join(lines), encoding="utf-8")
    console.print(f"[green]Markdown exported: {filename}[/green]")
    return filepath
