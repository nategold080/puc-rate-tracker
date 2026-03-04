"""EPA eGRID downloader and parser.

Downloads eGRID Excel workbooks containing utility-level emissions data:
CO2/NOx/SO2 emissions, emission rates, and generation mix.

Data source: https://www.epa.gov/egrid/download-data
Format: Multi-sheet Excel workbook
Key sheet: UTNL (Utility-level aggregation)
"""

from __future__ import annotations

import io
import time
from pathlib import Path
from typing import Any, Optional

import httpx
from rich.console import Console

console = Console()

PROJECT_ROOT = Path(__file__).parent.parent.parent
CACHE_DIR = PROJECT_ROOT / "data" / "raw" / "egrid"

# eGRID download URLs — these shift across EPA's site
EGRID_URLS = {
    2022: "https://www.epa.gov/system/files/documents/2024-01/egrid2022_data.xlsx",
    2021: "https://www.epa.gov/system/files/documents/2023-01/egrid2021_data.xlsx",
    2020: "https://www.epa.gov/system/files/documents/2022-01/egrid2020_data.xlsx",
}

USER_AGENT = (
    "PUC-Rate-Tracker/1.0 (Data research project; "
    "contact: nathanmauricegoldberg@gmail.com)"
)

# eGRID UTNL sheet column mappings
UTNL_COLUMNS = {
    "UTLSRVNM": "utility_name_egrid",
    "UTLSRVST": "state",
    "UTLSRVID": "eia_utility_id",
    "UTNGENAN": "net_generation_mwh",
    "UTLCO2AN": "co2_tons",
    "UTLNOXAN": "nox_tons",
    "UTLSO2AN": "so2_tons",
    "UTLCO2RA": "co2_rate_lbs_mwh",
    "UTLNOXRA": "nox_rate_lbs_mwh",
    "UTLSO2RA": "so2_rate_lbs_mwh",
    "UTLCLPR": "coal_pct",
    "UTLGSPR": "gas_pct",
    "UTLNCPR": "nuclear_pct",
    "UTLHYPR": "hydro_pct",
    "UTLWNPR": "wind_pct",
    "UTLSOPR": "solar_pct",
    "UTLOIPR": "other_renewable_pct",
}


def download_egrid(year: int, force: bool = False) -> Optional[Path]:
    """Download eGRID Excel file for a given year.

    Args:
        year: Data year (e.g., 2022).
        force: Re-download even if cached.

    Returns:
        Path to the downloaded file, or None on failure.
    """
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    file_path = CACHE_DIR / f"egrid{year}_data.xlsx"

    if file_path.exists() and not force:
        console.print(f"[dim]Using cached eGRID for {year}[/dim]")
        return file_path

    url = EGRID_URLS.get(year)
    if not url:
        console.print(f"[yellow]No known eGRID URL for {year}[/yellow]")
        return None

    console.print(f"[blue]Downloading eGRID {year}...[/blue]")

    try:
        with httpx.Client(
            follow_redirects=True, timeout=120.0,
            headers={"User-Agent": USER_AGENT}
        ) as client:
            resp = client.get(url)
            resp.raise_for_status()

        file_path.write_bytes(resp.content)
        console.print(f"[green]Downloaded {len(resp.content) / 1024 / 1024:.1f} MB[/green]")
        return file_path

    except Exception as e:
        console.print(f"[red]Error downloading eGRID {year}: {e}[/red]")
        return None


def parse_egrid(file_path: Path, year: int) -> list[dict[str, Any]]:
    """Parse eGRID utility-level emissions from Excel.

    Args:
        file_path: Path to the eGRID Excel file.
        year: Data year.

    Returns:
        List of emissions record dicts.
    """
    import openpyxl

    console.print(f"[dim]Parsing eGRID {year}...[/dim]")

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

    # Find UTNL sheet (utility-level, NOT unit-level UNT)
    utnl_sheet = None
    for name in wb.sheetnames:
        if "utnl" in name.lower():
            utnl_sheet = name
            break

    if not utnl_sheet:
        # Try other naming patterns
        for name in wb.sheetnames:
            if "utility" in name.lower():
                utnl_sheet = name
                break

    if not utnl_sheet:
        console.print(f"[red]Could not find utility sheet in eGRID {year}[/red]")
        console.print(f"[dim]Available sheets: {wb.sheetnames}[/dim]")
        wb.close()
        return []

    ws = wb[utnl_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Find header row
    header_idx = 0
    for i, row in enumerate(rows):
        row_vals = [str(c).strip() if c else "" for c in row]
        if "UTLSRVNM" in row_vals or "UTNGENAN" in row_vals:
            header_idx = i
            break

    headers = [str(c).strip() if c else "" for c in rows[header_idx]]

    # Map columns
    col_indices = {}
    for egrid_col, our_field in UTNL_COLUMNS.items():
        if egrid_col in headers:
            col_indices[our_field] = headers.index(egrid_col)

    if "utility_name_egrid" not in col_indices:
        console.print(f"[red]Missing utility name column in eGRID sheet[/red]")
        return []

    records = []
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        name_idx = col_indices["utility_name_egrid"]
        if name_idx >= len(row) or not row[name_idx]:
            continue

        rec = {"year": year}
        for field, idx in col_indices.items():
            if idx >= len(row):
                continue
            val = row[idx]
            if val is None:
                continue

            if field in ("utility_name_egrid", "state"):
                rec[field] = str(val).strip()
            elif field == "eia_utility_id":
                try:
                    rec[field] = int(float(str(val).strip()))
                except (ValueError, TypeError):
                    pass
            else:
                try:
                    rec[field] = round(float(str(val).strip()), 4)
                except (ValueError, TypeError):
                    pass

        # Quality score
        rec["quality_score"] = _score_emissions(rec)
        records.append(rec)

    console.print(f"[green]Parsed {len(records)} utility emissions records for {year}[/green]")
    return records


def _score_emissions(rec: dict) -> float:
    """Score quality of an emissions record."""
    score = 0.0
    if rec.get("utility_name_egrid"):
        score += 0.15
    if rec.get("state"):
        score += 0.10
    if rec.get("net_generation_mwh"):
        score += 0.15
    if rec.get("co2_tons"):
        score += 0.15
    if rec.get("co2_rate_lbs_mwh"):
        score += 0.15
    if rec.get("coal_pct") is not None:
        score += 0.10
    if rec.get("gas_pct") is not None:
        score += 0.10
    if rec.get("eia_utility_id"):
        score += 0.10
    return round(score, 3)


def fetch_egrid(
    years: Optional[list[int]] = None,
    force: bool = False,
) -> list[dict[str, Any]]:
    """Download and parse eGRID data for specified years.

    Args:
        years: List of years. Defaults to [2022].
        force: Re-download even if cached.

    Returns:
        Combined list of emissions records.
    """
    if years is None:
        years = [2022]

    all_records = []
    for year in years:
        file_path = download_egrid(year, force=force)
        if file_path:
            records = parse_egrid(file_path, year)
            all_records.extend(records)
            time.sleep(2)

    console.print(f"[bold green]Total eGRID records: {len(all_records)}[/bold green]")
    return all_records
